from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


def _style_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def export_to_excel(
    revenue_val: float,
    orders_val: int,
    avg_ticket_val: float,
    customers_val: int,
    top_products: pd.DataFrame,
) -> bytes:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Resumen"
    ws_summary.append(["Métrica", "Valor"])
    ws_summary.append(["Revenue Total", revenue_val])
    ws_summary.append(["Órdenes Totales", orders_val])
    ws_summary.append(["Ticket Promedio", avg_ticket_val])
    ws_summary.append(["Clientes Activos", customers_val])
    _style_header(ws_summary)

    ws_products = wb.create_sheet("Top Productos")
    for r_idx, row in enumerate(dataframe_to_rows(top_products, index=False, header=True)):
        ws_products.append(row)
    _style_header(ws_products)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
